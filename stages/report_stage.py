# file: stages/report_stage.py
"""
REPORT stage
report.source 설정에 따라 소스 DB 또는 target DB에서 SQL 실행 → CSV 저장 → Excel 변환.
skip_sql: true 시 DB 연결 없이 기존 CSV를 바로 union → Excel 처리.

job.yml 설정:
  report:
    source: target          # target(기본) / oracle / vertica
    skip_sql: false         # true: DB 연결 없이 csv_union_dir의 CSV만 사용
    csv_union_dir: data/export  # skip_sql=true 시 union할 CSV 소스 폴더
    export_csv:
      enabled: true
      sql_dir: sql/report/
      out_dir: data/report/
      compression: none     # none(기본) / gzip
    excel:
      enabled: true
      out_dir: data/report/
      max_files: 10
"""

import csv
import gzip
import re
import time
from datetime import datetime
from pathlib import Path

from engine.connection import connect_target
from engine.context import RunContext
from engine.path_utils import resolve_path
from engine.sql_utils import sort_sql_files, render_sql


def run(ctx: RunContext):
    logger = ctx.logger
    # logger.info("REPORT stage start")

    report_cfg = ctx.job_config.get("report")
    if not report_cfg:
        logger.info("REPORT stage skipped (no config)")
        return

    if ctx.mode == "plan":
        logger.info("REPORT stage skipped (plan mode)")
        return

    generated_csvs = []

    skip_sql = bool(report_cfg.get("skip_sql", False))

    if skip_sql:
        # DB 연결 없이 csv_union_dir 의 CSV 파일들을 바로 사용
        csv_union_dir = report_cfg.get("csv_union_dir", "data/export")
        union_dir = resolve_path(ctx, csv_union_dir)
        if union_dir.exists():
            generated_csvs = sorted(
                p for p in union_dir.rglob("*")
                if p.is_file() and (p.name.endswith(".csv") or p.name.endswith(".csv.gz"))
            )
            logger.info("REPORT skip_sql=true | csv_union_dir=%s files=%d",
                        union_dir, len(generated_csvs))
        else:
            logger.warning("REPORT skip_sql=true | csv_union_dir not found: %s", union_dir)
    else:
        export_csv_cfg = report_cfg.get("export_csv", {})
        if export_csv_cfg.get("enabled", False):
            generated_csvs = _run_csv_export(ctx, report_cfg, export_csv_cfg)
        else:
            logger.info("REPORT csv export skipped")

    excel_cfg = report_cfg.get("excel", {})
    if excel_cfg.get("enabled", False):
        _run_excel_export(ctx, report_cfg, excel_cfg, generated_csvs)
    else:
        logger.info("REPORT excel export skipped")

    # logger.info("REPORT stage end")


# ────────────────────────────────────────────────────────────
# CSV Export
# ────────────────────────────────────────────────────────────

def _run_csv_export(ctx, report_cfg, cfg) -> list:
    logger = ctx.logger
    sql_dir = resolve_path(ctx, cfg.get("sql_dir", "sql/report"))
    out_dir  = resolve_path(ctx, cfg.get("out_dir",  "data/report"))
    compression = (cfg.get("compression") or "none").strip().lower()

    out_dir.mkdir(parents=True, exist_ok=True)

    if not sql_dir.exists():
        logger.warning("REPORT sql_dir not found: %s", sql_dir)
        return []

    sql_files = sort_sql_files(sql_dir)
    if not sql_files:
        logger.info("REPORT no SQL files in %s", sql_dir)
        return []

    # report.source 결정: 기본값은 "target"
    report_source = (report_cfg.get("source") or "target").strip().lower()
    conn, conn_type, conn_label = _open_connection(ctx, report_source)

    logger.info(
        "REPORT csv export | db=%s sql_dir=%s out_dir=%s files=%d compression=%s",
        conn_label, sql_dir, out_dir, len(sql_files), compression,
    )

    generated = []
    total = len(sql_files)

    try:
        for i, sql_file in enumerate(sql_files, 1):
            sql_text = sql_file.read_text(encoding="utf-8")
            rendered = render_sql(sql_text, ctx.params)

            ext = ".csv.gz" if compression == "gzip" else ".csv"
            out_file = out_dir / (sql_file.stem + ext)

            logger.info("REPORT [%d/%d] %s → %s", i, total, sql_file.name, out_file.name)
            start = time.time()
            try:
                rows = _export_to_csv(conn, conn_type, rendered, out_file, compression)
                logger.info("REPORT [%d/%d] done | rows=%d elapsed=%.2fs", i, total, rows, time.time() - start)
                generated.append(out_file)
            except Exception as e:
                logger.error("REPORT [%d/%d] FAILED (%.2fs): %s", i, total, time.time() - start, e)
    finally:
        conn.close()

    return generated


def _open_connection(ctx, report_source: str):
    """
    report_source 에 따라 연결 반환.
      "target"          → job의 target DB (DuckDB/SQLite3/Oracle MYDATA)
      "oracle"/"vertica" → 소스 DB
    반환: (conn, conn_type, label)   실패 시 (None, None, None)
    """
    logger = ctx.logger

    if report_source == "target":
        target_cfg = ctx.job_config.get("target", {})
        return connect_target(ctx, target_cfg)

    # source DB (oracle / vertica)
    from adapters.sources.oracle_client import init_oracle_client, get_oracle_conn
    from adapters.sources.vertica_client import get_vertica_conn

    source_sel = ctx.job_config.get("source", {})
    src_type  = source_sel.get("type", "oracle")
    host_name = source_sel.get("host", "")
    env_cfg   = ctx.env_config

    if src_type == "oracle":
        oracle_cfg = env_cfg["sources"]["oracle"]
        host_cfg = oracle_cfg["hosts"].get(host_name)
        if not host_cfg:
            raise RuntimeError(f"Oracle host not found: {host_name}")
        init_oracle_client(oracle_cfg)
        dsn = host_cfg.get("dsn") or f"{host_cfg.get('host')}:{host_cfg.get('port', 1521)}/{host_cfg.get('service_name', '')}"
        label = f"oracle source ({dsn})"
        return get_oracle_conn(host_cfg), "oracle", label

    elif src_type == "vertica":
        vertica_cfg = env_cfg["sources"]["vertica"]
        host_cfg = vertica_cfg["hosts"].get(host_name)
        if not host_cfg:
            raise RuntimeError(f"Vertica host not found: {host_name}")
        label = f"vertica ({host_name})"
        return get_vertica_conn(host_cfg), "vertica", label

    else:
        raise ValueError(f"REPORT: unsupported source type: {src_type}")


def _export_to_csv(conn, conn_type: str, sql_text: str, out_file: Path, compression: str) -> int:
    """SQL 실행 결과를 CSV 저장. row 수 반환."""
    open_fn = gzip.open if compression == "gzip" else open
    row_count = 0

    if conn_type == "duckdb":
        rel = conn.execute(sql_text)
        columns = [d[0] for d in rel.description]
        with open_fn(out_file, "wt", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            while True:
                batch = rel.fetchmany(10000)
                if not batch:
                    break
                for row in batch:
                    writer.writerow(["" if v is None else str(v) for v in row])
                    row_count += 1
    else:
        cur = conn.cursor()
        try:
            if conn_type == "oracle":
                cur.arraysize = 10000
            cur.execute(sql_text)
            columns = [d[0] for d in cur.description]
            with open_fn(out_file, "wt", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                while True:
                    batch = cur.fetchmany(10000)
                    if not batch:
                        break
                    for row in batch:
                        writer.writerow(["" if v is None else str(v) for v in row])
                        row_count += 1
        finally:
            cur.close()

    return row_count


# ────────────────────────────────────────────────────────────
# Excel Export (pandas 기반, csv_to_excel 로직)
# ────────────────────────────────────────────────────────────

def _get_excel_output_path(ctx, out_dir: Path, job_name: str, max_files: int) -> Path:
    logger = ctx.logger
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%y%m%d")
    base = f"{job_name}_{ts}"
    pattern = re.compile(rf"{re.escape(base)}_(\d+)\.xlsx")

    existing = []
    for f in out_dir.glob(f"{base}_*.xlsx"):
        m = pattern.fullmatch(f.name)
        if m:
            existing.append((int(m.group(1)), f))

    next_idx = max((i for i, _ in existing), default=0) + 1
    out = out_dir / f"{base}_{next_idx}.xlsx"

    existing_sorted = sorted(existing, key=lambda x: x[1].stat().st_mtime)
    while len(existing_sorted) >= max_files:
        _, old = existing_sorted.pop(0)
        old.unlink()
        logger.info("REPORT excel: deleting old file %s", old.name)

    return out


def _run_excel_export(ctx, report_cfg, cfg, csv_files: list):
    logger = ctx.logger
    out_dir_str = cfg.get("out_dir") or report_cfg.get("export_csv", {}).get("out_dir", "data/report")
    out_dir = resolve_path(ctx, out_dir_str)
    max_files = int(cfg.get("max_files", 10))

    if not csv_files:
        if out_dir.exists():
            csv_files = sorted(
                p for p in out_dir.iterdir()
                if p.is_file() and (p.name.endswith(".csv") or p.name.endswith(".csv.gz"))
            )

    if not csv_files:
        logger.warning("REPORT excel: no target CSV found, skip")
        return

    output_path = _get_excel_output_path(ctx, out_dir, ctx.job_name, max_files)
    logger.info("REPORT excel export | output=%s files=%d", output_path.name, len(csv_files))
    start = time.time()

    try:
        import pandas as pd
        from pandas.api.types import is_integer_dtype, is_float_dtype
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        summary_rows = []

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

            for csv_file in csv_files:
                sheet_name = csv_file.stem.replace(".csv", "").upper()[:31]

                open_fn = gzip.open if str(csv_file).endswith(".gz") else open

                # 행 수 사전 체크 (OOM 방지)
                with open_fn(csv_file, "rt", encoding="utf-8") as f:
                    row_count = sum(1 for _ in f) - 1  # 헤더 제외
                if row_count > 1_048_576:
                    logger.warning("REPORT excel: row limit exceeded, skip | %s rows=%d", sheet_name, row_count)
                    continue

                with open_fn(csv_file, "rt", encoding="utf-8") as f:
                    df = pd.read_csv(f)

                df.to_excel(writer, sheet_name=sheet_name, index=False)
                summary_rows.append({"sheet_name": sheet_name, "rows": row_count})

                ws = writer.book[sheet_name]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                # 헤더 스타일
                header_fill = PatternFill("solid", fgColor="D9E1F2")
                header_font = Font(bold=True)
                for col_idx in range(1, len(df.columns) + 1):
                    c = ws.cell(row=1, column=col_idx)
                    c.fill = header_fill
                    c.font = header_font

                # 컬럼 너비
                for col_idx, col_name in enumerate(df.columns, start=1):
                    col_series = df.iloc[:, col_idx - 1].astype(str)
                    max_len = max(col_series.str.len().max(), len(str(col_name)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(int(max_len * 1.2) + 2, 50)


                # 숫자 컬럼 포맷
                # for col_idx, col_name in enumerate(df.columns, start=1):
                #     if is_integer_dtype(df[col_name]) or is_float_dtype(df[col_name]):
                #         for row in range(2, ws.max_row + 1):
                #             ws.cell(row=row, column=col_idx).number_format = "#,##0"

                logger.info("REPORT excel sheet: %s (%d rows)", sheet_name, row_count)

            # SUMMARY 시트
            if summary_rows:
                import pandas as pd
                summary_df = pd.DataFrame(summary_rows)
                summary_df.insert(0, "no", range(1, len(summary_df) + 1))
                summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

                ws_sum = writer.book["SUMMARY"]
                ws_sum.freeze_panes = "A2"
                ws_sum.auto_filter.ref = ws_sum.dimensions

                h_fill = PatternFill("solid", fgColor="BDD7EE")
                h_font = Font(bold=True)
                for col_idx in range(1, summary_df.shape[1] + 1):
                    c = ws_sum.cell(row=1, column=col_idx)
                    c.fill = h_fill
                    c.font = h_font

                for col_idx, col_name in enumerate(summary_df.columns, start=1):
                    values = [len(str(col_name))]
                    values += [len(str(v)) for v in summary_df[col_name] if pd.notna(v)]
                    max_len = max(values)
                    ws_sum.column_dimensions[get_column_letter(col_idx)].width = min(int(max_len * 1.2) + 2, 30)

                existing_sheets = set(writer.book.sheetnames)
                for row_idx in range(2, ws_sum.max_row + 1):
                    sheet_val = ws_sum.cell(row=row_idx, column=2).value
                    if sheet_val and sheet_val in existing_sheets:
                        ws_sum.cell(row=row_idx, column=2).hyperlink = f"#'{sheet_val}'!A1"
                        ws_sum.cell(row=row_idx, column=2).style = "Hyperlink"

                wb = writer.book
                wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws_sum)))

        logger.info("REPORT excel done | %s (%.2fs)", output_path.name, time.time() - start)

    except ImportError as e:
        logger.error("REPORT excel: package not installed (%s) -> pip install pandas openpyxl", e)
    except Exception as e:
        logger.exception("REPORT excel generation failed: %s", e)
